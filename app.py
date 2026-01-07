"""
Keyword Finder Application
A tool to mine long-tail keyword opportunities using Google's Autocomplete suggestions.
"""

import json
import re
import time
import csv
import io
import os
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, send_from_directory, Response, stream_with_context
from flask_cors import CORS
import requests
from dotenv import load_dotenv
from typing import List, Dict, Set, Tuple
from collections import Counter
import re

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app)

# Configuration from environment variables
GOOGLE_SUGGEST_URL = "http://google.com/complete/search"
REQUEST_DELAY = float(os.getenv('GOOGLE_REQUEST_DELAY', '0.3'))
MAX_PARALLEL_REQUESTS = 5

# Proxy Configuration from environment variables
GOOGLE_HTTP_PROXY = os.getenv('GOOGLE_HTTP_PROXY', '')
GOOGLE_HTTPS_PROXY = os.getenv('GOOGLE_HTTPS_PROXY', '')
GOOGLE_PROXIES_ROTATION = os.getenv('GOOGLE_PROXIES_ROTATION', '')


def get_proxies(use_proxy: bool = True) -> Dict[str, str]:
    """
    Get proxy configuration for requests.
    Returns a dict with 'http' and 'https' keys, or empty dict if no proxy configured or use_proxy is False.

    Priority:
    1. proxies.txt file (if exists) - format: IP:PORT:USERNAME:PASSWORD
    2. GOOGLE_PROXIES_ROTATION env var (comma-separated)
    3. GOOGLE_HTTP_PROXY and GOOGLE_HTTPS_PROXY env vars

    Args:
        use_proxy: If False, returns empty dict regardless of settings
    """
    if not use_proxy:
        return {}

    # First, try to read from proxies.txt file
    proxies_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'proxies.txt')
    if os.path.exists(proxies_file):
        try:
            with open(proxies_file, 'r') as f:
                proxy_lines = [line.strip() for line in f if line.strip()]
                if proxy_lines:
                    # Random proxy rotation from proxies.txt
                    selected_proxy = random.choice(proxy_lines)
                    # Parse format: IP:PORT:USERNAME:PASSWORD
                    parts = selected_proxy.split(':')
                    if len(parts) == 4:
                        ip, port, username, password = parts
                        proxy_url = f"http://{username}:{password}@{ip}:{port}"
                        return {
                            'http': proxy_url,
                            'https': proxy_url
                        }
        except Exception as e:
            print(f"Error reading proxies.txt: {e}")

    # If rotation list is configured, use a random proxy from the list
    if GOOGLE_PROXIES_ROTATION:
        proxy_list = [p.strip() for p in GOOGLE_PROXIES_ROTATION.split(',') if p.strip()]
        if proxy_list:
            selected_proxy = random.choice(proxy_list)
            return {
                'http': selected_proxy,
                'https': selected_proxy
            }

    # Otherwise, use individual HTTP/HTTPS proxy settings
    proxies = {}
    if GOOGLE_HTTP_PROXY:
        proxies['http'] = GOOGLE_HTTP_PROXY
    if GOOGLE_HTTPS_PROXY:
        proxies['https'] = GOOGLE_HTTPS_PROXY

    return proxies


class KeywordHarvester:
    """Main class for harvesting keywords from Google Autocomplete."""

    def __init__(self, use_proxy=True):
        self.use_proxy = use_proxy  # Whether to use proxy for requests
        self.modifiers = {
            'alphabet': list('abcdefghijklmnopqrstuvwxyz'),
            'numbers': list('0123456789'),
            'prepositions': ['for', 'with', 'without', 'in', 'near', 'at', 'on'],
            'comparisons': ['vs', 'or', 'versus', 'like', 'similar to'],
            'intents': ['best', 'top', 'cheap', 'review', 'guide', 'free'],
            'questions': ['how to', 'what is', 'why', 'where', 'can', 'does', 'will']
        }

    def fetch_suggestions_google(self, query: str, max_retries: int = 20, timeout: int = 10) -> tuple:
        """
        Fetch suggestions from Google Autocomplete API (free) with retry mechanism.

        Returns:
            tuple: (suggestions_list, retry_count, success)
        """
        import time as time_module

        for attempt in range(max_retries):
            try:
                params = {
                    'client': 'chrome',
                    'q': query,
                    'hl': 'en'
                }
                # Get proxy configuration based on use_proxy setting
                proxies = get_proxies(self.use_proxy)

                response = requests.get(
                    GOOGLE_SUGGEST_URL,
                    params=params,
                    proxies=proxies if proxies else None,
                    timeout=timeout
                )
                if response.status_code == 200:
                    data = response.json()
                    if len(data) > 1 and isinstance(data[1], list):
                        return data[1], attempt, True
                    else:
                        print(f"Unexpected response format for '{query}': {data[:100] if isinstance(data, list) else str(data)[:100]}")
                else:
                    print(f"Non-200 status for '{query}': {response.status_code}")
            except requests.exceptions.Timeout:
                print(f"Timeout (attempt {attempt + 1}/{max_retries}) fetching Google suggestions for '{query}'")
                if attempt < max_retries - 1:
                    # Exponential backoff: 2^attempt seconds, max 30 seconds
                    backoff = min(2 ** attempt, 30)
                    time_module.sleep(backoff)
            except requests.exceptions.RequestException as e:
                print(f"Request error (attempt {attempt + 1}/{max_retries}) for '{query}': {e}")
                if attempt < max_retries - 1:
                    backoff = min(2 ** attempt, 30)
                    time_module.sleep(backoff)
            except Exception as e:
                print(f"Unexpected error (attempt {attempt + 1}/{max_retries}) for '{query}': {e}")
                if attempt < max_retries - 1:
                    backoff = min(2 ** attempt, 30)
                    time_module.sleep(backoff)

        print(f"Failed to fetch suggestions for '{query}' after {max_retries} attempts")
        return [], max_retries, False

    def fetch_suggestions(self, query: str, **kwargs) -> tuple:
        """
        Fetch suggestions from Google Autocomplete with retry mechanism.

        Returns:
            tuple: (suggestions_list, retry_count, success)
        """
        return self.fetch_suggestions_google(query)

    def generate_queries(self, seed: str, selected_modifiers: Dict[str, bool]) -> List[str]:
        """Generate query variations based on selected modifiers."""
        queries = set()

        # A-Z alphabet soup
        if selected_modifiers.get('alphabet'):
            for letter in self.modifiers['alphabet']:
                queries.add(f"{seed} {letter}")

        # Numbers
        if selected_modifiers.get('numbers'):
            for num in self.modifiers['numbers']:
                queries.add(f"{seed} {num}")

        # Prepositions
        if selected_modifiers.get('prepositions'):
            for prep in self.modifiers['prepositions']:
                queries.add(f"{seed} {prep}")

        # Comparisons
        if selected_modifiers.get('comparisons'):
            for comp in self.modifiers['comparisons']:
                queries.add(f"{seed} {comp}")

        # Intent/Qualifiers (prepend and append)
        if selected_modifiers.get('intents'):
            for intent in self.modifiers['intents']:
                queries.add(f"{intent} {seed}")
                queries.add(f"{seed} {intent}")

        # Questions (prepend)
        if selected_modifiers.get('questions'):
            for q in self.modifiers['questions']:
                queries.add(f"{q} {seed}")

        # Always include the seed itself
        queries.add(seed)

        return list(queries)

    def harvest(self, seed: str, selected_modifiers: Dict[str, bool],
                progress_callback=None, **kwargs) -> List[Dict[str, str]]:
        """
        Harvest keywords using the selected modifiers.
        Returns a list of dicts with keyword and source information.
        """
        queries = self.generate_queries(seed, selected_modifiers)
        results = []
        seen_keywords = set()

        for i, query in enumerate(queries):
            if progress_callback:
                progress_callback(i + 1, len(queries), len(results), query, seed)

            suggestions, _, _ = self.fetch_suggestions(query)

            for suggestion in suggestions:
                # Clean and normalize
                clean_kw = self.clean_keyword(suggestion)

                if clean_kw and clean_kw not in seen_keywords:
                    seen_keywords.add(clean_kw)
                    results.append({
                        'keyword': clean_kw,
                        'source': query,
                        'seed': seed,
                        'char_count': len(clean_kw),
                        'word_count': len(clean_kw.split())
                    })

            # Rate limiting
            if i < len(queries) - 1:
                time.sleep(REQUEST_DELAY)

        return results

    @staticmethod
    def clean_keyword(keyword: str) -> str:
        """Clean and normalize a keyword."""
        # Remove extra whitespace
        keyword = ' '.join(keyword.split())

        # Convert to lowercase for normalization
        keyword = keyword.lower()

        # Remove non-ASCII characters (keep basic alphanumeric and punctuation)
        keyword = re.sub(r'[^\x00-\x7F]+', '', keyword)

        return keyword.strip()

    @staticmethod
    def filter_keywords(keywords: List[Dict], filters: Dict) -> List[Dict]:
        """Apply filters to the keyword list."""
        filtered = keywords

        # Include filter
        if filters.get('include'):
            include_text = filters['include'].lower()
            filtered = [k for k in filtered if include_text in k['keyword'].lower()]

        # Exclude filter
        if filters.get('exclude'):
            exclude_terms = [e.strip().lower() for e in filters['exclude'].split(',')]
            filtered = [k for k in filtered if
                       not any(term in k['keyword'].lower() for term in exclude_terms)]

        # Word count filters
        if filters.get('min_words', 0) > 0:
            filtered = [k for k in filtered if k['word_count'] >= filters['min_words']]

        if filters.get('max_words', 999) < 999:
            filtered = [k for k in filtered if k['word_count'] <= filters['max_words']]

        return filtered

    @staticmethod
    def extract_ngrams(keywords: List[Dict], top_n: int = 15) -> List[Dict]:
        """
        Extract single-word topics from keywords using unigrams.
        Returns a list of topics with their frequency counts.

        Args:
            keywords: List of keyword dictionaries
            top_n: Number of top topics to return
        """
        if not keywords:
            return []

        # Common stopwords to exclude from topics
        stopwords = {
            'a', 'an', 'the', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
            'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
            'should', 'may', 'might', 'must', 'can', 'for', 'of', 'with', 'at',
            'by', 'from', 'in', 'on', 'to', 'into', 'onto', 'upon', 'as',
            'or', 'and', 'but', 'if', 'then', 'when', 'where', 'how', 'what',
            'which', 'who', 'whom', 'whose', 'why', 'whether', 'while', 'after',
            'before', 'until', 'unless', 'since', 'because', 'although', 'though',
            'get', 'got', 'make', 'made', 'take', 'use', 'used', 'best', 'good',
            'new', 'old', 'more', 'most', 'some', 'such', 'than', 'too', 'very',
            'just', 'also', 'like', 'well', 'out', 'up', 'via', 'vs', 'versus'
        }

        # Extract all single words (unigrams)
        word_counts = Counter()

        for kw in keywords:
            words = kw['keyword'].lower().split()
            for word in words:
                # Skip stopwords and very short words
                if len(word) > 2 and word not in stopwords:
                    word_counts[word] += 1

        # Get top words by frequency
        top_words = word_counts.most_common(top_n)

        # Build result - just topics with counts, no pre-assigned keywords
        result = []
        for topic_word, count in top_words:
            # Calculate actual keyword count (keywords containing this word)
            keyword_count = sum(1 for kw in keywords if topic_word in kw['keyword'].lower())

            result.append({
                'topic': topic_word,
                'count': count,
                'keyword_count': keyword_count
            })

        return result


# Global harvester instance
harvester = KeywordHarvester()


@app.route('/')
def index():
    """Serve the main application page."""
    return send_from_directory('.', 'index.html')


def process_single_query(harvester_instance, query, seed, seen_keywords, results_lock):
    """Process a single query and return results (thread-safe)."""
    try:
        suggestions, retry_count, success = harvester_instance.fetch_suggestions(query)
        new_keywords = []

        with results_lock:
            for suggestion in suggestions:
                clean_kw = harvester_instance.clean_keyword(suggestion)
                if clean_kw and clean_kw not in seen_keywords:
                    seen_keywords.add(clean_kw)
                    keyword_data = {
                        'keyword': clean_kw,
                        'source': query,
                        'seed': seed,
                        'char_count': len(clean_kw),
                        'word_count': len(clean_kw.split())
                    }
                    new_keywords.append(keyword_data)

        return {
            'success': success,
            'query': query,
            'seed': seed,
            'keywords': new_keywords,
            'retries': retry_count
        }
    except Exception as e:
        return {
            'success': False,
            'query': query,
            'seed': seed,
            'error': str(e),
            'retries': 0
        }


def generate_sse(harvester_instance, seeds: List[str], modifiers: Dict[str, bool], workers: int = 1, **kwargs):
    """Generate Server-Sent Events for progress updates with multi-threading support."""
    import sys

    # Get number of workers from kwargs or default to 1
    num_workers = max(1, min(workers, 20))  # Limit to max 20 workers

    # First, calculate total queries across all seeds for accurate progress tracking
    all_seed_queries = []
    for seed in seeds:
        queries = harvester_instance.generate_queries(seed, modifiers)
        all_seed_queries.append((seed, queries))

    total_queries_all_seeds = sum(len(queries) for _, queries in all_seed_queries)

    # Send initial total to frontend
    init_data = {
        'type': 'init',
        'total_queries': total_queries_all_seeds,
        'total_seeds': len(seeds),
        'workers': num_workers
    }
    yield f"data: {json.dumps(init_data)}\n\n"

    # Thread-safe data structures
    all_results = []
    seen_keywords = set()
    results_lock = threading.Lock()
    completed_count = {'value': 0}
    count_lock = threading.Lock()

    try:
        # Flatten all queries into a list with metadata
        all_queries = []
        for seed_idx, (seed, queries) in enumerate(all_seed_queries):
            for query in queries:
                all_queries.append((query, seed, seed_idx, len(queries)))

        # Process queries in parallel using ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            # Submit all queries to the executor
            future_to_query = {
                executor.submit(process_single_query, harvester_instance, query, seed, seen_keywords, results_lock): (query, seed, seed_idx)
                for query, seed, seed_idx, _ in all_queries
            }

            # Process results as they complete
            seed_keyword_counts = {}  # Track keywords per seed
            total_retries = {'value': 0}  # Track total retries
            retry_lock = threading.Lock()

            for future in as_completed(future_to_query):
                query, seed, seed_idx = future_to_query[future]

                try:
                    result = future.result(timeout=60)  # 60 second timeout per query

                    # Track retries
                    with retry_lock:
                        total_retries['value'] += result.get('retries', 0)

                    with count_lock:
                        completed_count['value'] += 1
                        current = completed_count['value']

                    # Send progress update
                    progress_data = {
                        'type': 'progress',
                        'current': current,
                        'total': total_queries_all_seeds,
                        'found': len(all_results),
                        'query': query,
                        'seed': seed,
                        'seed_index': seed_idx,
                        'total_seeds': len(seeds),
                        'retries': result.get('retries', 0),
                        'total_retries': total_retries['value']
                    }
                    yield f"data: {json.dumps(progress_data)}\n\n"
                    sys.stdout.flush()

                    if result['success'] and result['keywords']:
                        with results_lock:
                            all_results.extend(result['keywords'])
                            # Track per-seed counts
                            if seed not in seed_keyword_counts:
                                seed_keyword_counts[seed] = 0
                            seed_keyword_counts[seed] += len(result['keywords'])

                        # Send new keywords in real-time
                        keywords_data = {
                            'type': 'keywords',
                            'keywords': result['keywords'],
                            'found': len(all_results),
                            'query': query,
                            'seed': seed,
                            'retries': result.get('retries', 0)
                        }
                        yield f"data: {json.dumps(keywords_data)}\n\n"
                        sys.stdout.flush()

                    # Rate limiting (reduced delay with threading)
                    time.sleep(REQUEST_DELAY / num_workers)

                except Exception as e:
                    print(f"Error processing query '{query}': {e}")

        # Send completion events for each seed
        for seed_idx, (seed, _) in enumerate(all_seed_queries):
            complete_data = {
                'type': 'seed_complete',
                'seed': seed,
                'count': seed_keyword_counts.get(seed, 0),
                'seed_index': seed_idx,
                'total_seeds': len(seeds)
            }
            yield f"data: {json.dumps(complete_data)}\n\n"

        # Send final completion event
        final_data = {
            'type': 'complete',
            'total': len(all_results),
            'keywords': all_results
        }
        yield f"data: {json.dumps(final_data)}\n\n"
    except GeneratorExit:
        # Client disconnected
        print("Client disconnected, stopping harvest")
    except Exception as e:
        print(f"Error during harvest: {e}")
        error_data = {
            'type': 'error',
            'message': str(e)
        }
        yield f"data: {json.dumps(error_data)}\n\n"


@app.route('/api/harvest', methods=['POST'])
def harvest_keywords():
    """API endpoint to harvest keywords with streaming progress."""
    data = request.json

    # Get seeds - support both single seed and bulk seeds
    seeds_input = data.get('seeds', '') or data.get('seed', '')
    if isinstance(seeds_input, str):
        seeds = [s.strip() for s in seeds_input.split('\n') if s.strip()]
    else:
        seeds = seeds_input

    if not seeds:
        return jsonify({'error': 'At least one seed keyword is required'}), 400

    selected_modifiers = {
        'alphabet': data.get('alphabet', True),
        'numbers': data.get('numbers', False),
        'prepositions': data.get('prepositions', False),
        'comparisons': data.get('comparisons', True),
        'intents': data.get('intents', True),
        'questions': data.get('questions', True)
    }

    # Proxy setting
    use_proxy = data.get('use_proxy', True)

    # Create harvester with proxy setting
    harvester_instance = KeywordHarvester(use_proxy=use_proxy)

    # Get number of workers (default to 1 for single-threaded)
    workers = data.get('workers', 1)

    return Response(
        stream_with_context(generate_sse(harvester_instance, seeds, selected_modifiers, workers=workers)),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'
        }
    )


@app.route('/api/filter', methods=['POST'])
def filter_keywords():
    """API endpoint to filter harvested keywords."""
    data = request.json
    keywords = data.get('keywords', [])

    filters = {
        'include': data.get('include', ''),
        'exclude': data.get('exclude', ''),
        'min_words': data.get('min_words', 0),
        'max_words': data.get('max_words', 999)
    }

    filtered = harvester.filter_keywords(keywords, filters)

    return jsonify({
        'success': True,
        'keywords': filtered,
        'total': len(filtered)
    })


@app.route('/api/topics', methods=['POST'])
def extract_topics():
    """API endpoint to extract n-gram based topics from keywords."""
    data = request.json
    keywords = data.get('keywords', [])
    top_n = data.get('top_n', 15)

    if not keywords:
        return jsonify({
            'success': True,
            'topics': [],
            'total': 0
        })

    topics = KeywordHarvester.extract_ngrams(keywords, top_n)

    return jsonify({
        'success': True,
        'topics': topics,
        'total': len(topics)
    })


@app.route('/api/export/csv', methods=['POST'])
def export_csv():
    """Export keywords to CSV format."""
    data = request.json
    keywords = data.get('keywords', [])

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['Keyword', 'Source', 'Character Count', 'Word Count'])

    # Data rows
    for kw in keywords:
        writer.writerow([
            kw.get('keyword', ''),
            kw.get('source', ''),
            kw.get('char_count', 0),
            kw.get('word_count', 0)
        ])

    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': 'attachment; filename=keywords.csv'
    }


@app.route('/api/export/txt', methods=['POST'])
def export_txt():
    """Export keywords to plain text (one per line)."""
    data = request.json
    keywords = data.get('keywords', [])

    lines = [kw.get('keyword', '') for kw in keywords]
    output = '\n'.join(lines)

    return output, 200, {
        'Content-Type': 'text/plain',
        'Content-Disposition': 'attachment; filename=keywords.txt'
    }


@app.route('/api/import', methods=['POST'])
def import_keywords():
    """Import keywords from uploaded file (CSV, TXT, or Excel)."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    keywords = []
    filename = file.filename.lower()

    try:
        if filename.endswith('.txt'):
            # Read as plain text (one keyword per line)
            content = file.read().decode('utf-8')
            for line in content.split('\n'):
                keyword = line.strip()
                if keyword:
                    # Clean and normalize
                    clean_kw = KeywordHarvester.clean_keyword(keyword)
                    if clean_kw:
                        keywords.append({
                            'keyword': clean_kw,
                            'source': 'import',
                            'seed': 'imported',
                            'char_count': len(clean_kw),
                            'word_count': len(clean_kw.split())
                        })

        elif filename.endswith('.csv'):
            # Read CSV file
            import io
            content = file.read().decode('utf-8')
            csv_reader = csv.reader(io.StringIO(content))

            # Skip header if it exists
            first_row = next(csv_reader, None)
            if first_row and first_row[0].lower() not in ['keyword', 'keywords', 'term', 'query']:
                # First row is not a header, process it
                keyword = first_row[0].strip()
                if keyword:
                    clean_kw = KeywordHarvester.clean_keyword(keyword)
                    if clean_kw:
                        keywords.append({
                            'keyword': clean_kw,
                            'source': 'import',
                            'seed': 'imported',
                            'char_count': len(clean_kw),
                            'word_count': len(clean_kw.split())
                        })

            # Process remaining rows
            for row in csv_reader:
                if row:  # Skip empty rows
                    keyword = row[0].strip()
                    if keyword:
                        clean_kw = KeywordHarvester.clean_keyword(keyword)
                        if clean_kw:
                            keywords.append({
                                'keyword': clean_kw,
                                'source': 'import',
                                'seed': 'imported',
                                'char_count': len(clean_kw),
                                'word_count': len(clean_kw.split())
                            })

        elif filename.endswith(('.xlsx', '.xls')):
            # Read Excel file
            import openpyxl
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Start from row 1, skip if it looks like a header
            start_row = 0
            first_cell = sheet.cell(row=1, column=1).value
            if first_cell and str(first_cell).lower() in ['keyword', 'keywords', 'term', 'query']:
                start_row = 2  # Skip header row

            for row in range(start_row, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value:
                    keyword = str(cell_value).strip()
                    if keyword:
                        clean_kw = KeywordHarvester.clean_keyword(keyword)
                        if clean_kw:
                            keywords.append({
                                'keyword': clean_kw,
                                'source': 'import',
                                'seed': 'imported',
                                'char_count': len(clean_kw),
                                'word_count': len(clean_kw.split())
                            })
        else:
            return jsonify({'success': False, 'error': 'Unsupported file format. Use .txt, .csv, or .xlsx'}), 400

        return jsonify({
            'success': True,
            'keywords': keywords,
            'total': len(keywords)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'Error importing file: {str(e)}'}), 500


def get_libraries_file_path():
    """Get the path to the libraries.json file."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'libraries.json')


def load_libraries():
    """Load libraries from the JSON file."""
    libraries_file = get_libraries_file_path()
    if os.path.exists(libraries_file):
        try:
            with open(libraries_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading libraries: {e}")
            return {}
    return {}


def save_libraries(libraries):
    """Save libraries to the JSON file."""
    libraries_file = get_libraries_file_path()
    try:
        with open(libraries_file, 'w', encoding='utf-8') as f:
            json.dump(libraries, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving libraries: {e}")
        return False


@app.route('/api/libraries', methods=['GET'])
def get_libraries():
    """Get all libraries."""
    libraries = load_libraries()
    return jsonify({
        'success': True,
        'libraries': libraries
    })


@app.route('/api/libraries', methods=['POST'])
def create_or_update_library():
    """Create or update a library."""
    data = request.json
    library_name = data.get('name', '').strip()
    words = data.get('words', [])

    if not library_name:
        return jsonify({'success': False, 'error': 'Library name is required'}), 400

    # Ensure words is a list and filter empty strings
    if isinstance(words, str):
        words = [w.strip() for w in words.split('\n') if w.strip()]
    else:
        words = [w.strip() for w in words if w.strip()]

    libraries = load_libraries()

    # Store library with words
    libraries[library_name] = words

    if save_libraries(libraries):
        return jsonify({
            'success': True,
            'library': library_name,
            'words': words
        })
    else:
        return jsonify({'success': False, 'error': 'Failed to save library'}), 500


@app.route('/api/libraries/<library_name>', methods=['DELETE'])
def delete_library(library_name):
    """Delete a library."""
    libraries = load_libraries()

    if library_name not in libraries:
        return jsonify({'success': False, 'error': 'Library not found'}), 404

    del libraries[library_name]

    if save_libraries(libraries):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Failed to delete library'}), 500


if __name__ == '__main__':
    print("Starting Keyword Finder Application...")
    print("Open your browser to http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
